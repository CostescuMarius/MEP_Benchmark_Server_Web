import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import json
from response_time_multi_threading import average_response_time_for_GET, average_response_time_for_POST
from output_rate import incremental_output_rate_test
import threading
import requests
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime


####### EXCEL #######
HISTORY_FILE = "History.xlsx"

def init_history_file():
    if not os.path.exists(HISTORY_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Teste"
        ws.append(["Data", "URL", "Metoda", "Tip Test", "Numar Requesturi", "Rezultat"])
        wb.save(HISTORY_FILE)


def log_to_excel(url, method, test_type, num_requests, result_summary):
    wb = load_workbook(HISTORY_FILE)
    ws = wb["Teste"]

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([
        now,
        url,
        method,
        test_type,
        num_requests if num_requests is not None else "-",
        result_summary.strip()
    ])

    wb.save(HISTORY_FILE)





def run_test():
    progress.start(10)
    run_button.config(state='disabled')


    url = url_entry.get()
    method = method_combo.get()
    test_type = test_type_combo.get()

    # setare numar de cereri in caz de test tip "Timp de raspuns"
    try:
        if test_type == "Rata de iesire":
            num_requests = None
        elif test_type == "Timp de raspuns":
            num_requests = int(requests_entry.get())
            if num_requests <= 0:
                raise ValueError
    except ValueError:
        messagebox.showerror("Eroare", "Numarul de cereri trebuie sa fie un numar pozitiv.")
        return


    # setare continut cerere in cazde cerere tip "POST"
    if method == "POST":
        try:
            data = json.loads(payload_text.get("1.0", tk.END).strip())
        except json.JSONDecodeError:
            messagebox.showerror("Eroare", "JSON invalid.")
            return
    else:
        data = None

    result_text.config(state='normal')
    result_text.delete("1.0", tk.END)

    if test_type == "Timp de raspuns":
        if method == "POST":
            avg, min_, max_ = average_response_time_for_POST(url, data, num_requests)
        else:
            avg, min_, max_ = average_response_time_for_GET(url, num_requests)

        if avg is not None:
            result_text.insert(tk.END, f"Timp mediu: {avg:.2f} ms\n")
            result_text.insert(tk.END, f"Timp minim: {min_:.2f} ms\n")
            result_text.insert(tk.END, f"Timp maxim: {max_:.2f} ms\n")

    elif test_type == "Rata de iesire":
        results = incremental_output_rate_test(url, result_text, method, data)
        max_rate = 0
        for count, rate in results:
            max_rate = max(max_rate, rate)
            #result_text.insert(tk.END, f"{count} requesturi => {rate:.2f} req/s\n")
        result_text.insert(tk.END, f"Output Rate: {max_rate:.2f} req/s\n")


    result_summary = result_text.get("1.0", tk.END)
    log_to_excel(url, method, test_type, num_requests, result_summary)
    result_text.config(state='disabled')

    progress.stop()
    run_button.config(state='normal')



def toggle_testType_input(*args):
    if test_type_combo.get() == "Timp de raspuns":
        requests_label.grid()
        requests_entry.grid()
    elif test_type_combo.get() == "Rata de iesire":
        requests_label.grid_remove()
        requests_entry.grid_remove()



def toggle_requestType_input(*args):
    if method_combo.get() == "POST":
        payload_label.grid()
        payload_text.grid()
    elif method_combo.get() == "GET":
        payload_label.grid_remove()
        payload_text.grid_remove()


def valid_url(validity, reason):
    if reason == 0: # empty
        url_status_label.config(text="❌ EMPTY URL", fg="red")
        run_button.config(state='disabled')
    
    elif reason == 1: # validity
        if not validity:
            url_status_label.config(text="❌ INVALID URL", fg="red")
            run_button.config(state='disabled')
        else:
            url_status_label.config(text="✔️ URL VALID", fg="green")
            run_button.config(state='normal')


def check_url():
    url = url_entry.get().strip()
    if not url:
        valid_url(False, 0)
        return
    
    try:
        #response = requests.head(url, timeout=2)
        response = requests.get(url, timeout=2, stream=True)
        if response.status_code < 400:
            valid_url(True, 1)
            
        else:
            valid_url(False, 1)
    except:
        valid_url(False, 1)


def on_url_change(event):
    threading.Thread(target=check_url, daemon=True).start()




# GUI 
init_history_file()
root = tk.Tk()
root.title("Benchmark Server Web")


tk.Label(root, text="URL:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
url_entry = tk.Entry(root, width=75)
url_entry.grid(row=0, column=1, sticky="w", padx=5, pady=5)
url_entry.bind("<KeyRelease>", on_url_change)
url_status_label = tk.Label(root, text="", font=("Arial", 10, "bold"))
url_status_label.grid(row=0, column=2,  sticky="w", padx=5, pady=5)


tk.Label(root, text="Tip Test:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
test_type_combo = ttk.Combobox(root, values=["Timp de raspuns", "Rata de iesire"], state="readonly")
test_type_combo.grid(row=1, column=1, sticky="w", padx=5, pady=5)
test_type_combo.current(0)
test_type_combo.bind("<<ComboboxSelected>>", toggle_testType_input)


tk.Label(root, text="Metoda:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
method_combo = ttk.Combobox(root, values=["GET", "POST"], state="readonly")
method_combo.grid(row=2, column=1, sticky="w", padx=5, pady=5)
method_combo.current(0)
method_combo.bind("<<ComboboxSelected>>", toggle_requestType_input)


requests_label = tk.Label(root, text="Numar requesturi:")
requests_label.grid(row=3, column=0, sticky="w", padx=5, pady=5)
requests_entry = tk.Entry(root)
requests_entry.insert(0, "10")
requests_entry.grid(row=3, column=1, sticky="w", padx=5, pady=5)


payload_label = tk.Label(root, text="Date JSON (POST):")
payload_label.grid(row=4, column=0, sticky="w", padx=5, pady=5)
payload_text = scrolledtext.ScrolledText(root, height=5, width=40)
payload_text.grid(row=4, column=1, sticky="w", padx=5, pady=5)
payload_label.grid_remove()
payload_text.grid_remove()


#run_button = tk.Button(root, text="Rulare test", command=run_test)
def threaded_run():
    threading.Thread(target=run_test, daemon=True).start()
run_button = tk.Button(root, text="Rulare test", command=threaded_run)
run_button.grid(row=5, column=0, columnspan=3, pady=10)
valid_url(False, 0)

result_text = scrolledtext.ScrolledText(root, height=10, width=60, state='disabled')
result_text.grid(row=6, column=0, columnspan=3, padx=10, pady=5)

progress = ttk.Progressbar(root, orient="horizontal", mode="indeterminate", length=400)
progress.grid(row=7, column=0, columnspan=3, padx=10, pady=10)

root.mainloop()
